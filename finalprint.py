import tkinter as tk
from tkinter import ttk
from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from tkinter import *
import openpyxl as xl
import math
Year = []
root = tk.Tk()
root.title('Financial Year Selector')
root.resizable(True, True)
root.geometry('250x150')

Label(root, text="Enter the Financial Year", font=('Calibri 10')).pack()
a=Entry(root, width=35)
a.pack()

def yeargetter(Year):
    Year.append( int(a.get()[0:4]))
    Year.append(Year[0] + 1)
    Year[0] = str(Year[0])
    Year[1] = str(Year[1])
    root.destroy()

def Yeargetf():
    yeargetter(Year)
    
ttk.Button(root, text="Next", command= Yeargetf).pack()

root.mainloop()
# create the root window
root = tk.Tk()
root.title('Input File Selector')
root.resizable(True, True)
root.geometry('550x650')

prevdir= [-1]


temp = [tk.StringVar(),tk.StringVar(),tk.StringVar(),tk.StringVar()]

temp[0].set("Opening Balances")
temp[1].set("Outstanding Loan");
temp[2].set("Receipts");
temp[3].set("Payments");


def select_files(i):
    filetypes = (
        ('All files', '*.*'),
        ('text files', '*.txt')
        
    )
    if (prevdir[0] == -1):
      di = '/'
    else:
      di = prevdir[0]
    
    filenames = fd.askopenfilenames(
       
        title='Open files',
        initialdir=di,# holds the key to pop up recent use one
        filetypes=filetypes)
    
    h = temp[i].get()
    
    prevdir[0] = h[2:-3]
    temp[i].set( filenames)
    
    if (temp[i].get() == ''):
      temp[i].set( h)
    
def select_files2(i):
   
    filenames = fd.askdirectory(
        title='Open files',
        initialdir='/',
        )
    h = temp[i].get()
    temp[i].set( filenames)
    

   
    if (temp[i].get() == ''):
      temp[i].set( h)
    #showinfo(

     #   title='Selected Files',
      #  message=filenames
    #)
def f1():
    select_files(0)
def f2():
    select_files(1)
def f3():
    select_files(2)
def f4():
    select_files(3)

def f5():
    select_files2(0)

f = [f1,f2,f3,f4]
# open button
open_button = [ttk.Button(
    root,
    textvariable = temp[i],
    command=f[i]
) for i in range(4)]




open_button[0].pack(expand=True)
open_button[1].pack(expand=True)
open_button[2].pack(expand=True)
open_button[3].pack(expand=True)

ttk.Button(root, text="Next", command=root.destroy).pack()

root.mainloop()


    
wb_obj = [xl.load_workbook(temp[i].get()[2:-3]) for i in range(4)]
sheet_obj = [wb_obj[i].active for i in range(4)]



def dategetter(l):
    l = str(l)
    k = l[8:10] + l[4:7] + l[4:5] + l[0:4]
    return k

class Row:

    def __init__(self):
        self.ReceiptNo = ""
        self.Date = "01-04-"+ Year[0]
        self.Village = ""
        self.GroupNo = ""
        self.Membership = ""
        self.Savings = ""
        self.ShareCapital = ""
        self.DRA = ""
        self.Principal = ""
        self.Interest = ""
        self.Stationary = ""
        

    def loan(self,sheet,r):
        self.Date = sheet.cell(row = r,column = 1).value
        
        self.Village = sheet.cell(row = r,column = 2).value
        self.GroupNo = sheet.cell(row = r,column = 3).value
        temp = sheet.cell(row = r,column = 5).value
        if (temp != None):
            self.Principal = '-' + str(temp)
        self.DRA = sheet.cell(row = r ,column = 6).value

    def Opening(self,sheet,r):
        
        self.GroupNo = sheet.cell(row = r , column = 5).value
        self.Savings = sheet.cell(row = r,column = 7).value
        self.ShareCapital = sheet.cell(row = r,column = 8).value

    def loanOutStanding(self,sheet,r):
        self.GroupNo = sheet.cell(row = r,column = 5).value
        self.Principal = -1*int(sheet.cell(row = r,column = 7).value)
        self.Interest = -1*int(sheet.cell(row = r,column = 12).value)

    def writeto(self,sheet,r):
        sheet.cell(row = r,column = 1).value = self.ReceiptNo
        h = sheet.cell(row = r,column = 2)
        h.value = self.Date
            
        sheet.cell(row = r,column = 3).value = self.Village
        sheet.cell(row = r,column = 4).value = self.GroupNo
        sheet.cell(row = r,column = 5).value = self.Membership
        sheet.cell(row = r,column = 6).value = self.Savings
        sheet.cell(row = r,column = 7).value = self.ShareCapital
        sheet.cell(row = r,column = 8).value = self.DRA
        sheet.cell(row = r,column = 9).value = self.Principal
        sheet.cell(row = r,column = 10).value = self.Interest
        sheet.cell(row = r,column = 11).value = self.Stationary
        

def addingloans(fromsheet,tosheet):
    
    n = tosheet.max_row
    
   
    for i in range(2,fromsheet.max_row + 1):
        x  = Row()
        x.loan(fromsheet,i)
        n += 1
        x.writeto(tosheet,n)

        tosheet.cell(row = n ,column = 2).value = dategetter(tosheet.cell(row = n,column = 2).value)
       


def addingopen(fromsheet,tosheet):
    
    n = tosheet.max_row
    
    for i in range(2, fromsheet.max_row + 1):
        x  = Row()
        x.Opening(fromsheet,i)
        n += 1
        x.writeto(tosheet,n)
        #tosheet.cell(row = n ,column = 2).value = dategetter(tosheet.cell(row = n,column = 2).value)

def addingoutstan(fromsheet,tosheet):
    n = tosheet.max_row
    
    for i in range(2,fromsheet.max_row + 1):
        x = Row()
        x.loanOutStanding(fromsheet,i)
        n += 1
        x.writeto(tosheet,n)
        #tosheet.cell(row = n ,column = 2).value = dategetter(tosheet.cell(row = n,column = 2).value)

def adding(sheet_obj):

   
    addingloans(sheet_obj[3],sheet_obj[2])
    addingopen(sheet_obj[0],sheet_obj[2])
    addingoutstan(sheet_obj[1],sheet_obj[2])


sheet_obj2= wb_obj[2].active

for i in range(2,sheet_obj2.max_row + 1 ):
    h =sheet_obj2.cell(row = i,column = 2)
    
    h.value = dategetter(h.value)

#wb_obj[2].save("firstem.xlsx")
adding(sheet_obj)
#wb_obj[2].save("merged.xlsx")

monthdays = [0,31,28,31,30,31,30,31,31,30,31,30,31]

for i in range(1,13):
    monthdays[i] += monthdays[i - 1]
    
#dd/mm/yyyy
def pre(date)->int:
   

     #days till the date from start of year
   day = date[0: 2]
   month = date[3 : 5]

   
   
   
   year = date[6:10]
   
  
   year = int(year)
   m = int(month)

   su = 0
   su += monthdays[m - 1]
   su += int(day)
   leap = 0
   if (year % 4 == 0 ):
       leap = 1

   if (m > 2):
       su += leap
       
   return su


def suff(date)->int:      #days after the date till the end of year
    year = int(date[6:10])

    leap = 0

    if (year % 4 ==0 and year % 400 != 0):
       leap = 1

    return (365) + leap - pre(date)

def daydiffere(start,end)->int:
    start = str(start)
    end= str(end)
    
    year1 = start[6:10]
    year2 = end[6:10]

    if (year1 == year2):
      return pre(end) - pre(start)

    return suff(start) + pre(end)
#r1 > r2 true else false
def cmp(wb,r1,r2):
    date1 = wb.cell(row = r1,column = 2).value
    date2 = wb.cell(row = r2,column = 2).value
     # 0:2  3:5 6:10
    year1 = date1[6 : 10]
    year2 = date2[6 : 10]

    if (year1 != year2):
        return year1 > year2

    month1 = date1[3 : 5]
    month2 = date2[3 : 5]

    if (month1 != month2):
      return month1 > month2

    day1 = date1[0:2]
    day2 = date2[0:2]

    return day1 > day2

    

   
def writer(fromshe,toshe,r1,r2):
    for i in range(1,12):
        toshe.cell(row = r2,column = i).value = fromshe.cell(row = r1,column = i).value


        
#wb is the original sheet
#wb2 is the auxilary sheet
def sorter(wb,wb2,l,r):

    if (l == r):
      return
    
    mid = (l + r)//2

    sorter(wb,wb2,l,mid)
    sorter(wb,wb2,mid + 1,r)

    t = l
    t1 = l
    t2 = mid + 1
    while (t1 <= mid and t2 <= r):
    

        if (cmp(wb,t2,t1) == True):
           writer(wb,wb2,t1,t)
           t1 += 1
           t += 1
        else:
            writer(wb,wb2,t2,t)
            t2 += 1
            t += 1



    while (t1 <= mid):
           writer(wb,wb2,t1,t)
           t1 += 1
           t += 1


    while (t2 <= r):
            writer(wb,wb2,t2,t)
            t2 += 1
            t += 1


    t = l
    while (t <= r):
          writer(wb2,wb,t,t)
          t += 1

dummy_work = xl.Workbook()
dummy_sheet = dummy_work.active




 

sorter(sheet_obj[2],dummy_sheet,  2, sheet_obj[2].max_row)



Interest_on_savings = 4
Interest_on_ShareCapital = 4
Interest_on_Principal = 24

Start_Date = "01-04-"+ Year[0]

class Account:

    def __init__(self):
        self.id = - 1
        self.savings = 0
        self.savingsinterest = 0
        self.savingdate = Start_Date
        self.savingstransactions = []
        
        self.ShareCapital = 0
        self.ShareCapitalinterest = 0
        self.ShareCapitaldate = Start_Date
        self.ShareCapitaltransactions = []

        self.Principal = 0
        self.Principalinterest = 0
        self.Principaldate = Start_Date
        self.Principaltransactions = []

        self.Membersh = 0
        self.Membershipdate = Start_Date
        self.Membershiptransactions = []
       
        self.DRA = 0
        self.DRAdate = Start_Date
        self.DRAtransactions = []

        self.Stationary = 0
        self.Stationarydate = Start_Date
        self.Stationarytransactions = []

    def  Membershipamount(self,amount,date):
         if (amount != None and amount != ''):
             amount = int(amount)
             self.Membersh += amount
             self.Membershiptransactions.append((date,amount))
             self.Membershipdate = date

    def  DRAamount(self,amount,date):
         if (amount != None and amount != ''):
             amount = int(amount)
             self.DRA += amount
             self.DRAtransactions.append((date,amount))
             self.DRAdate = date

    def  Stationaryamount(self,amount,date):
         if (amount != None and amount != ''):
             amount = int (amount)
             self.Stationary += amount
             self.Stationarytransactions.append((date,amount))
             self.Stationarydate = date

    def Savingsamount(self,amount,date):
         if (amount != None and amount != ''):
             amount = int (amount)
             days = daydiffere(self.savingdate,date)
             self.savingdate = date
             h = (self.savings*Interest_on_savings*days)/(36500)
             h = math.floor(h)
             self.savings += amount
             self.savingsinterest += h
             self.savingstransactions.append((date,amount,h))
             


    def ShareCapitalamount(self,amount,date):
         if (amount != None and amount != ''):
             amount = int (amount)
             days = daydiffere(self.ShareCapitaldate,date)
             self.ShareCapitaldate = date
             h = (self.ShareCapital*Interest_on_ShareCapital*days)/(36500)
             h = math.floor(h)
             self.ShareCapital += amount
             self.ShareCapitalinterest += h
             self.ShareCapitaltransactions.append((date,amount,h))
            

    def Principalamount(self,amount,interest,date):
         if (amount != None and amount != ''):
             amount = int (amount)
             days = daydiffere(self.Principaldate,date)
             self.Principaldate = date
             h = (self.Principal*Interest_on_Principal*days)/(36500)
             #print(h,int(h),days)
             h = math.ceil(h)
             if (interest == None or interest == ''):
                 k = 0
             else:
                 k = int(interest)
             if (k >= 0):
             #k is interest received and h is interest receivable
                 self.Principaltransactions.append((date,amount,k,h))
                 
             else:
                 self.Principaltransactions.append((date,amount,0,-k))

             self.Principal -= amount
             self.Principalinterest += h
                 
             
         

gp_name_mapper = {}
gp_name_list = []


gAccounts = []
processingsheet = sheet_obj[2]
wb_obj[2].save("prefinale.xlsx")
#groupno column 4
for i in range(2,processingsheet.max_row + 1):
    gp_name = processingsheet.cell(row = i,column = 4).value
    if (gp_name in gp_name_mapper):
         gid = gp_name_mapper[gp_name]
    else:
        gid = len(gp_name_mapper)
        gp_name_mapper[gp_name] = gid
        gp_name_list.append(gp_name)
        gAccounts.append(Account())
    #print(gp_name,"group-name")
    gac = gAccounts[gid]
    date = processingsheet.cell(row = i,column  = 2).value
    gac.Membershipamount(processingsheet.cell(row = i,column  = 5).value,date)
    gac.DRAamount(processingsheet.cell(row = i,column  = 8).value,date)
    gac.Stationaryamount(processingsheet.cell(row = i,column  = 11).value,date)
    gac.Savingsamount(processingsheet.cell(row = i,column  = 6).value,date)
    gac.ShareCapitalamount(processingsheet.cell(row = i,column  = 7).value,date)
    gac.Principalamount(processingsheet.cell(row = i,column  = 9).value,processingsheet.cell(row = i,column  = 10).value,date)
    
    
enddate = "31-03-"+ Year[1]
gp_name_list.sort()


for i in range(len(gp_name_list)):
    gp_name = gp_name_list[i]
    gid = gp_name_mapper[gp_name]
    gac = gAccounts[gid]

    date = enddate
    gac.Membershipamount(0,date)
    gac.DRAamount(0,date)
    gac.Stationaryamount(0,date)
    gac.Savingsamount(0,date)
    gac.ShareCapitalamount(0,date)
    gac.Principalamount(0,0,date)
    
        

        
wb = xl.Workbook()


ws1 = wb.create_sheet(title = "Group Wise Annual Statement")
ws2 = wb.create_sheet(title = "Annual Summary ")
## Date:GpName:Membership:DRA:Stationary:Savings:SavingsInterest:ShareCapital:Sharingcapitalinterest:Principal:Interestreceivable:Interestreceived
#never use one in naming coz l and one both are printed bloody same
ws1.cell(row = 1,column = 1).value = "DATE"
ws1.cell(row = 1,column = 2).value = "Group Name"
ws1.cell(row = 1,column = 3).value = "Membership"
ws1.cell(row = 1,column = 4).value = "DRA"
ws1.cell(row = 1,column = 5).value = "Stationary"
ws1.cell(row = 1,column = 6).value = "Savings"
ws1.cell(row = 1,column = 7).value = "SavingsInterest"
ws1.cell(row = 1,column = 8).value = "ShareCapital"
ws1.cell(row = 1,column = 9).value = "ShareCapitalInterest"
ws1.cell(row = 1,column = 10).value = "Principal"
ws1.cell(row = 1,column = 11).value = "Interestreceivable"
ws1.cell(row = 1,column = 12).value = "Interestreceived"


no_of_accounts = len(gp_name_list)
final_sum = []
row_no_output = 2
for i in range(no_of_accounts):
     group_name = gp_name_list[i]
     group_id = gp_name_mapper[group_name]
     ws1.cell(row = row_no_output,column = 2).value = group_name
     row_no_output += 1

     total = 0

     gacn = gAccounts[group_id]
     temp_sum = []
     su = 0

     for j in range(len(gacn.Membershiptransactions)):
        h = gacn.Membershiptransactions[j]
        ws1.cell(row = row_no_output,column = 1).value = h[0]
        ws1.cell(row = row_no_output,column = 3).value = h[1]
        su += h[1]
        row_no_output += 1
        
     ws1.cell(row = row_no_output,column = 2).value = "Total"
     ws1.cell(row = row_no_output,column = 3).value = su
     row_no_output += 1
     temp_sum.append(su)
     su = 0

     for j in range(len(gacn.DRAtransactions)):
         h= gacn.DRAtransactions[j]
         ws1.cell(row = row_no_output,column = 1).value = h[0]
         ws1.cell(row = row_no_output,column = 4).value = h[1]
         su += h[1]
         row_no_output += 1
                 
     ws1.cell(row = row_no_output,column = 2).value = "Total"
     ws1.cell(row = row_no_output,column = 4).value = su
     row_no_output += 1
     temp_sum.append(su)
     su = 0

     for j in range(len(gacn.Stationarytransactions)):
         h = gacn.Stationarytransactions[j]
         ws1.cell(row = row_no_output,column = 1).value = h[0]
         ws1.cell(row = row_no_output,column = 5).value = h[1]
         su += int(h[1])
         row_no_output += 1

     ws1.cell(row = row_no_output,column = 2).value = "Total"
     ws1.cell(row = row_no_output,column = 5).value = su
     row_no_output += 1
     temp_sum.append(su)
     su = 0
     su2 = 0

     for j in range(len(gacn.savingstransactions)):
         h = gacn.savingstransactions[j]
         ws1.cell(row = row_no_output,column = 1).value = h[0]
         ws1.cell(row = row_no_output,column = 6).value = h[1]
         ws1.cell(row = row_no_output,column = 7).value = h[2]
         su += int(h[1])
         su2 += int(h[2])
         row_no_output += 1


     ws1.cell(row = row_no_output,column = 2).value = "Total"
     ws1.cell(row = row_no_output,column = 6).value = su
     ws1.cell(row = row_no_output,column = 7).value = str(su2)
     row_no_output += 1
     temp_sum.append(su)
     temp_sum.append(su2)
     su = 0
     su2 = 0

     for j in range(len(gacn.ShareCapitaltransactions)):
         h = gacn.ShareCapitaltransactions[j]
         ws1.cell(row = row_no_output,column = 1).value = h[0]
         ws1.cell(row = row_no_output,column = 8).value = h[1]
         ws1.cell(row = row_no_output,column = 9).value = h[2]
         su += int(h[1])
         su2 += int(h[2])
         row_no_output += 1


     ws1.cell(row = row_no_output,column = 2).value = "Total"
     ws1.cell(row = row_no_output,column = 8).value = str(su)
     ws1.cell(row = row_no_output,column = 9).value = str(su2)
     row_no_output += 1
     temp_sum.append(su)
     temp_sum.append(su2)
     su = 0
     su2 = 0
     su3 = 0

     for j in range(len(gacn.Principaltransactions)):
         h = gacn.Principaltransactions[j]
         ws1.cell(row = row_no_output,column = 1).value = h[0]
         ws1.cell(row = row_no_output,column = 10).value = h[1]
         ws1.cell(row = row_no_output,column = 11).value = h[2]
         ws1.cell(row = row_no_output,column = 12).value = h[3]
         su += int(h[1])
         su2 += int(h[2])
         su3 += int(h[3])
         row_no_output += 1


     ws1.cell(row = row_no_output,column = 2).value = "Total"
     ws1.cell(row = row_no_output,column = 10).value = str(su)
     ws1.cell(row = row_no_output,column = 11).value = str(su2)
     ws1.cell(row = row_no_output,column = 12).value = str(su3)
     row_no_output += 1
     temp_sum.append(-su)
     temp_sum.append(su2)
     temp_sum.append(su3)

     row_no_output += 2


     final_sum.append(temp_sum)


ws2.cell(row = 1,column = 1).value = "S.No"
ws2.cell(row = 1,column = 2).value = "Group Name"
ws2.cell(row = 1,column = 3).value = "Membership"
ws2.cell(row = 1,column = 4).value = "DRA"
ws2.cell(row = 1,column = 5).value = "Stationary"
ws2.cell(row = 1,column = 6).value = "Savings"
ws2.cell(row = 1,column = 7).value = "SavingsInterest"
ws2.cell(row = 1,column = 8).value = "ShareCapital"
ws2.cell(row = 1,column = 9).value = "ShareCapitalInterest"
ws2.cell(row = 1,column = 10).value = "Principal"
ws2.cell(row = 1,column = 11).value = "Interestreceived"
ws2.cell(row = 1,column = 12).value = "Interestreceivable"

last_row = [0 for i in range(10)]

second_sheet_rows = 2
for i in range(len(gp_name_list)):

    ws2.cell(row = second_sheet_rows,column = 1).value = str(i + 1)
    ws2.cell(row = second_sheet_rows,column = 2).value = gp_name_list[i]

    for j in range(3,13):
        ws2.cell(row = second_sheet_rows,column = j).value = str(final_sum[i][j - 3])
        last_row[j - 3] += final_sum[i][j - 3]

    
    second_sheet_rows += 1




ws2.cell(row = second_sheet_rows,column = 1 ).value = "Total"


for j in range(3,13):
    ws2.cell(row = second_sheet_rows, column = j).value = str(last_row[j - 3])


    

root = tk.Tk()
root.title('Output file location Selector')
root.resizable(True, True)
root.geometry('550x650')

temp = [tk.StringVar()]

temp[0].set("Output folder");

open_button= ttk.Button(root, textvariable = temp[0], command = f5)
open_button.pack(expand = True)
ttk.Button(root, text="Next", command=root.destroy).pack()

root.mainloop()

finale = temp[0].get()[2:]

#wb = wb_obj[2]
wb.save(finale + "/AnnualAnalysis.xlsx")





